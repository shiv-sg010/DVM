import logging
from datetime import date
import openpyxl

class book():
#constructor to initialize class attributes like name,author,isbn
    def __init__(self,name,author,isbn):
            self.name=name
            self.author=author
            self.isbn=isbn
#method to find out the index value of any book in the log
    def index(list,bname):
        c=0
        for i in list:
            if len(i)>= 4:
                if i[5]!='':
                    c=c+1
                    continue
            if (i[0]==bname):
                return c
            c=c+1
#to issue a book to the user available in the library
    def issue(bname,uname):
        if shelf.ifexist(book.catalogue,bname)==True:
            shelf.library_init(bname)
            if book.library_log[book.index(book.library_log,bname)][4]!= '' or book.library_log[book.index(book.library_log,bname)][6]!='' : 
                print("book is not available right now ")
            else:
                logging.basicConfig(format='%(asctime)s - %(message)s',level=logging.INFO)
                logging.info(" book issued")
                book.library_log[book.index(book.library_log,bname)][3]=uname
                d=date.today()
                book.library_log[book.index(book.library_log,bname)][4]=" issued on " +d.strftime("%B %d, %Y")
        else:
            print("No such book available in library ")
#to return any book issued by the user
    def retrn(bname,uname):
            if shelf.ifexist(book.catalogue,bname)==True:
                if shelf.ifexist(book.library_log,bname)== True:
                    shelf.library_init(bname)
                    if  book.library_log[book.index(book.library_log,bname)][5]!= '':
                        print("book is already returned ")
                    elif book.library_log[book.index(book.library_log,bname)][4]=='' or book.library_log[book.index(book.library_log,bname)][3]!= uname:
                        print("No such book has been issued ")
                    else:
                        logging.basicConfig(format='%(asctime)s - %(message)s',level=logging.INFO)
                        logging.info("book returned ")
                        d=date.today()
                        book.library_log[book.index(book.library_log,bname)][5]= " returned on "+d.strftime("%B %d, %Y")
                else: print("No such book has been issued ")
                if shelf.library_log != []:
                    shelf.library_log.pop()      
            else:
                print("No such book available in library ")
#to reserve any book available in the library that hasnt been issues by the user
    def reserved(bname,uname):
        if shelf.ifexist(book.catalogue,bname)==True:
            if shelf.ifexist(book.library_log,bname)== False:
                    shelf.library_init(bname)
                    logging.basicConfig(format='%(asctime)s - %(message)s',level=logging.INFO)
                    logging.info("book reserved ")
                    book.library_log[book.index(book.library_log,bname)][6]= "reserved"
            else:
                if  book.library_log[book.index(book.library_log,bname)][4]!='' : 
                    print("book is already issued to someone else ")
                else:
                    logging.basicConfig(format='%(asctime)s - %(message)s',level=logging.INFO)
                    logging.info(" book reserved ")
                    book.library_log[book.index(book.library_log,bname)][6]= "reserved"
                    book.library_log[book.index(book.library_log,bname)][3]=uname
        else:
            print("No such book available in library ")
    catalogue=[]              # nested list to store all the books in excel file
    library_log=[]            #nested list containing log of all the books
class shelf(book):
#initialzes and store book details from the excel file 
    def __init__(self,name,author,isbn):
        super().__init__(name,author,isbn)
        l=[name,author,isbn]
        book.catalogue.append(l)
#check whether if a book exist in the library
    def ifexist(list,bname):
        for i in list:
            if i[0]==bname:
                return True
        return False
 #log the issue return reserves in library list
    def library_init(bname):
        if shelf.ifexist(book.catalogue,bname)== True:
            l=[]
            for i in book.catalogue[book.index(book.catalogue,bname)]:
                l.append(i)
            l.extend(['','','',''])
            book.library_log.append(l)
#print all the book stored in library with book name isbn and author name
    def show_catalogue():
           print(*('{} {} {}'.format(*r) for r in book.catalogue), sep='\n')
#print the logging details such as book issued book returned from the librarylogger
    def show_logger():
           for i in book.library_log:
                print(i[0],i[1],i[2],i[3]," status: ",i[4],i[5],i[6],"\n")
#to add any new book in the library
    def add():
        bname=input("enter book name ")        
        if shelf.ifexist(book.catalogue,bname)== False:
            isbn=(input("enter isbn number "))
            author=input("enter author name ")
            book.catalogue.append([bname,isbn,author])
        else:
            print("this book already exists ")
#to remove any book from the library
    def remove(bname):
        if shelf.ifexist(book.catalogue,bname)== True:
             del book.catalogue[book.index(book.catalogue,bname)]
             print("book deleted ")
        else:
            print("no sucn book in library ")
#to  print the count all the book in library
    def book_count():
        print(len(book.catalogue))

#extract all the book details from the excel file
    def populate_book():
        l=[]
        path="C:\\Users\\f2022\Desktop\\C C++\\.vscode\\Book1.xlsx"  #path of the excel sheet
        wb_obj= openpyxl.load_workbook(path)    
        sheet_obj= wb_obj.active 
        m=sheet_obj.max_row
        n=sheet_obj.max_column
        for i in range (1,m+1):
            for j in range(1,n+1):
                cell_obj=sheet_obj.cell(row = i, column=j)
                l.append(cell_obj.value)
            shelf(l[0],l[1],l[2])
            l=[]

class login(shelf):
    
    print("WELCOME ")
    shelf.populate_book()

    def call():
        print("Hey there ")
        print("Do you wanna login as user or librarian ")
        ch=input("u for user l for librarian \n")
        if ch=="u":
            uname=input("enter your username ")
            logging.basicConfig(format='%(asctime)s - %(message)s',level=logging.INFO)
            logging.info("logged in as user ")
            print("Welcome ",uname)
            print("what do wanna do today ")
#a user function that can issue or return or reserve any book
            def user():
                ch=input("press\n i --> to issue a book \n r --> to return a book\n rv -->to reserve a book\n back -->to go back to login page \n exit --> to exit the library \n")
                if ch=="i":
                    bname = input("enter the book name ")
                    book.issue(bname, uname)
                elif ch=="r":
                    bname=input("enter the book name ")
                    book.retrn(bname,uname)
                elif ch=="rv":
                    bname=input("enter the book name ")
                    book.reserved(bname,uname)
                elif ch=="back":
                     login.call()
                elif ch=="exit":
                     print("Thank You ")
                     exit()
                else:
                    logging.warning('Invalid input!!! ')
                user()
            user()

        elif ch=="l":
            uname=input("enter your name ")
            logging.basicConfig(format='%(asctime)s - %(message)s',level=logging.INFO)
            logging.info("logged in as librarian ")
            print("Welcome ",uname)
            print("what do wanna do today ")
#method to access all the book in library and make any changes (accessed by a librarian)
            def librarian():
                ch=input("enter \n show --> to see the book catalogue\n log --> to see logger\n add --> to add book\n remove --> to remove book \n count --> to book count\n back --> to go back to login page\n exit --> to exit the library \n")
                if ch=="show":
                    shelf.show_catalogue()
                elif ch=="log":
                    shelf.show_logger()
                elif ch=="add":
                    shelf.add()
                elif ch=="remove":
                    bname=input("enter the book name ")
                    shelf.remove(bname)
                elif ch=="count":
                    shelf.book_count()
                elif ch=="back":
                    login.call()
                elif ch=="exit":
                     print("Thank You ")
                     exit()
                else:
                    logging.warning('Invalid input!!! ')
                librarian()
            librarian()
        elif ch=="3":
            exit()
        else:
            logging.warning("Invalid input!!! ")
        login.call()

login.call()


    



